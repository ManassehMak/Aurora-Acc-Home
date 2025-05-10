import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import random
from datetime import datetime, timedelta

# Create workbook
wb = openpyxl.Workbook()

# Define formatting
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
cell_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
input_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

# Function to set column widths
def set_column_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

# Function to apply header formatting
def format_headers(ws, headers, row=1):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = cell_border

# Function to apply borders to range
def apply_borders(ws, start_row, start_col, end_row, end_col):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = cell_border

# Sheet: Справочники
ws_ref = wb.create_sheet("Справочники")
suppliers = [
    (1, "МБН-Маркет"), (2, "Петербургцемент"), (3, "ПУЛЬС"), (4, "ООО КВК"), (5, "Петрович")
]
materials = [
    (1, "Арматура д12", "кг"), (2, "Цемент М500 Д20", "кг"), (3, "Гранулы ПП", "м³"),
    (4, "СДО", "л"), (5, "Пластификатор ССМ-Р1", "л")
]
expense_types = [(1, "Аренда"), (2, "Зарплата"), (3, "Прочее")]

# Suppliers table
ws_ref.cell(row=1, column=1).value = "ID_Поставщика"
ws_ref.cell(row=1, column=2).value = "Название"
format_headers(ws_ref, ["ID_Поставщика", "Название"])
for i, (id_sup, name) in enumerate(suppliers, 2):
    ws_ref.cell(row=i, column=1).value = id_sup
    ws_ref.cell(row=i, column=2).value = name
apply_borders(ws_ref, 1, 1, 6, 2)

# Materials table
ws_ref.cell(row=1, column=4).value = "ID_Материала"
ws_ref.cell(row=1, column=5).value = "Название"
ws_ref.cell(row=1, column=6).value = "Единица_измерения"
format_headers(ws_ref, ["ID_Материала", "Название", "Единица_измерения"], row=1)
for i, (id_mat, name, unit) in enumerate(materials, 2):
    ws_ref.cell(row=i, column=4).value = id_mat
    ws_ref.cell(row=i, column=5).value = name
    ws_ref.cell(row=i, column=6).value = unit
apply_borders(ws_ref, 1, 4, 6, 6)

# Expense types table
ws_ref.cell(row=1, column=9).value = "ID_Типа"
ws_ref.cell(row=1, column=10).value = "Название"
format_headers(ws_ref, ["ID_Типа", "Название"], row=1)
for i, (id_type, name) in enumerate(expense_types, 2):
    ws_ref.cell(row=i, column=9).value = id_type
    ws_ref.cell(row=i, column=10).value = name
apply_borders(ws_ref, 1, 9, 4, 10)

set_column_widths(ws_ref, {1: 15, 2: 20, 4: 15, 5: 20, 6: 20, 9: 15, 10: 15})

# Sheet: Закупки
ws_purch = wb.create_sheet("Закупки")
purch_headers = ["ID_Закупки", "Дата", "Поставщик", "Материал", "Количество", "Сумма", "Год", "Квартал"]
format_headers(ws_purch, purch_headers)
set_column_widths(ws_purch, {1: 15, 2: 15, 3: 20, 4: 20, 5: 15, 6: 15, 7: 10, 8: 10})

# Generate random purchase data
base_date = datetime(2023, 1, 1)
purchases = []
for i in range(1, 6):
    date = base_date + timedelta(days=random.randint(0, 180))
    supplier = random.choice(suppliers)[1]
    material = random.choice(materials)
    quantity = random.randint(1000, 5000) if material[2] == "кг" else random.randint(1, 10) if material[2] == "м³" else random.randint(50, 200)
    price_per_unit = 100 if material[2] == "кг" else 60000 if material[2] == "м³" else 500
    amount = quantity * price_per_unit
    purchases.append((i, date, supplier, material[1], f"{quantity} {material[2]}", amount))

for i, (id_purch, date, supp, mat, qty, amt) in enumerate(purchases, 2):
    ws_purch.cell(row=i, column=1).value = id_purch
    ws_purch.cell(row=i, column=2).value = date
    ws_purch.cell(row=i, column=3).value = supp
    ws_purch.cell(row=i, column=4).value = mat
    ws_purch.cell(row=i, column=5).value = qty
    ws_purch.cell(row=i, column=6).value = amt
    ws_purch.cell(row=i, column=7).value = f"=YEAR(B{i})"
    ws_purch.cell(row=i, column=8).value = f"=CEILING(MONTH(B{i})/3,1)"
apply_borders(ws_purch, 1, 1, 6, 8)

# Data Validation for Закупки
dv_supp = DataValidation(type="list", formula1='=Справочники!$B$2:$B$6', allow_blank=True)
dv_supp.add("C2:C100")
ws_purch.add_data_validation(dv_supp)
dv_mat = DataValidation(type="list", formula1='=Справочники!$E$2:$E$6', allow_blank=True)
dv_mat.add("D2:D100")
ws_purch.add_data_validation(dv_mat)

# Sheet: Заливки
ws_pours = wb.create_sheet("Заливки")
pour_headers = ["ID_Заливки", "Дата", "ID_Сборки", "Объем_м3", "Стоимость_цемента", "Стоимость_гранул",
                "Стоимость_СДО", "Стоимость_арматуры", "Себестоимость_м3", "Расход_цемента_тн",
                "Расход_гранул_м3", "Расход_СДО_л", "Расход_арматуры_тн"]
format_headers(ws_pours, pour_headers)
set_column_widths(ws_pours, {1: 15, 2: 15, 3: 15, 4: 12, 5: 18, 6: 18, 7: 15, 8: 18, 9: 18, 10: 18, 11: 18, 12: 15, 13: 18})

# Generate random pour data
pours = []
for i in range(1, 6):
    date = base_date + timedelta(days=random.randint(0, 180))
    assembly_id = f"{date.strftime('%m/%d')}-{random.randint(1, 5)}"
    volume = round(random.uniform(4, 10), 2)
    cement_cost = random.randint(10000, 20000)
    granules_cost = random.randint(8000, 15000)
    sdo_cost = random.randint(100, 300)
    rebar_cost = random.randint(4000, 8000)
    cement_usage = round(volume * random.uniform(0.3, 0.4), 2)
    granules_usage = round(volume * random.uniform(0.5, 0.8), 2)
    sdo_usage = round(volume * random.uniform(0.2, 0.4), 2)
    rebar_usage = round(volume * random.uniform(0.01, 0.03), 3)
    pours.append((i, date, assembly_id, volume, cement_cost, granules_cost, sdo_cost, rebar_cost,
                  cement_usage, granules_usage, sdo_usage, rebar_usage))

for i, (id_pour, date, ass_id, vol, c_cost, g_cost, s_cost, r_cost, c_use, g_use, s_use, r_use) in enumerate(pours, 2):
    ws_pours.cell(row=i, column=1).value = id_pour
    ws_pours.cell(row=i, column=2).value = date
    ws_pours.cell(row=i, column=3).value = ass_id
    ws_pours.cell(row=i, column=4).value = vol
    ws_pours.cell(row=i, column=5).value = c_cost
    ws_pours.cell(row=i, column=6).value = g_cost
    ws_pours.cell(row=i, column=7).value = s_cost
    ws_pours.cell(row=i, column=8).value = r_cost
    ws_pours.cell(row=i, column=9).value = f"=(E{i}+F{i}+G{i}+H{i})/D{i}"
    ws_pours.cell(row=i, column=10).value = c_use
    ws_pours.cell(row=i, column=11).value = g_use
    ws_pours.cell(row=i, column=12).value = s_use
    ws_pours.cell(row=i, column=13).value = r_use
apply_borders(ws_pours, 1, 1, 6, 13)

# Sheet: Финансовые_расходы
ws_exp = wb.create_sheet("Финансовые_расходы")
exp_headers = ["ID_Расхода", "Дата", "Тип_расхода", "Сумма", "Вклад_в_себестоимость_м3"]
format_headers(ws_exp, exp_headers)
set_column_widths(ws_exp, {1: 15, 2: 15, 3: 15, 4: 15, 5: 25})

# Generate random expense data
expenses = []
for i in range(1, 6):
    date = base_date + timedelta(days=random.randint(0, 180))
    exp_type = random.choice(expense_types)[1]
    amount = random.randint(10000, 300000) if exp_type == "Прочее" else random.randint(250000, 600000)
    expenses.append((i, date, exp_type, amount))

for i, (id_exp, date, exp_type, amt) in enumerate(expenses, 2):
    ws_exp.cell(row=i, column=1).value = id_exp
    ws_exp.cell(row=i, column=2).value = date
    ws_exp.cell(row=i, column=3).value = exp_type
    ws_exp.cell(row=i, column=4).value = amt
    ws_exp.cell(row=i, column=5).value = f"=D{i}/SUM(Заливки!D:D)"
apply_borders(ws_exp, 1, 1, 6, 5)

# Data Validation for Финансовые_расходы
dv_exp = DataValidation(type="list", formula1='=Справочники!$J$2:$J$4', allow_blank=True)
dv_exp.add("C2:C100")
ws_exp.add_data_validation(dv_exp)

# Sheet: Ввод_данных
ws_form = wb.create_sheet("Ввод_данных")
# Form: Закупки
form_purch = ["Поле", "Значение", "Дата", "Поставщик", "Материал", "Количество", "Сумма"]
for i, label in enumerate(form_purch, 1):
    cell = ws_form.cell(row=i, column=1)
    cell.value = label
    cell.font = Font(bold=True)
    cell.border = cell_border
    if i > 2:
        ws_form.cell(row=i, column=2).fill = input_fill
        ws_form.cell(row=i, column=2).border = cell_border

dv_supp_form = DataValidation(type="list", formula1='=Справочники!$B$2:$B$6', allow_blank=True)
dv_supp_form.add("B4")
ws_form.add_data_validation(dv_supp_form)
dv_mat_form = DataValidation(type="list", formula1='=Справочники!$E$2:$E$6', allow_blank=True)
dv_mat_form.add("B5")
ws_form.add_data_validation(dv_mat_form)

# Form: Заливки
form_pour = ["Поле", "Значение", "Дата", "ID_Сборки", "Объем_м3", "Стоимость_цемента", "Стоимость_гранул",
             "Стоимость_СДО", "Стоимость_арматуры", "Расход_цемента_тн", "Расход_гранул_м3", "Расход_СДО_л",
             "Расход_арматуры_тн"]
for i, label in enumerate(form_pour, 1):
    cell = ws_form.cell(row=i, column=4)
    cell.value = label
    cell.font = Font(bold=True)
    cell.border = cell_border
    if i > 2:
        ws_form.cell(row=i, column=5).fill = input_fill
        ws_form.cell(row=i, column=5).border = cell_border

# Form: Расходы
form_exp = ["Поле", "Значение", "Дата", "Тип_расхода", "Сумма"]
for i, label in enumerate(form_exp, 1):
    cell = ws_form.cell(row=i, column=11)
    cell.value = label
    cell.font = Font(bold=True)
    cell.border = cell_border
    if i > 2:
        ws_form.cell(row=i, column=12).fill = input_fill
        ws_form.cell(row=i, column=12).border = cell_border

dv_exp_form = DataValidation(type="list", formula1='=Справочники!$J$2:$J$4', allow_blank=True)
dv_exp_form.add("L4")
ws_form.add_data_validation(dv_exp_form)

set_column_widths(ws_form, {1: 15, 2: 20, 4: 20, 5: 20, 11: 15, 12: 20})

# Sheet: Дашборд
ws_dash = wb.create_sheet("Дашборд")
ws_dash.cell(row=1, column=1).value = "Инструкции по настройке дашборда:"
ws_dash.cell(row=2, column=1).value = "1. Создайте сводную таблицу для Закупки (Источник: Закупки!A:H, Строки: Год, Квартал, Значения: Сумма, Количество)."
ws_dash.cell(row=3, column=1).value = "2. Создайте столбчатую диаграмму для Сумма по кварталам."
ws_dash.cell(row=4, column=1).value = "3. Создайте сводную таблицу для Заливки (Источник: Заливки!A:M, Строки: Месяц, Год, Значения: Объем_м3, Себестоимость_м3)."
ws_dash.cell(row=5, column=1).value = "4. Создайте линейный график для Себестоимость_м3 по месяцам."
ws_dash.cell(row=6, column=1).value = "5. Создайте сводную таблицу для Финансовые_расходы (Источник: Финансовые_расходы!A:E, Строки: Тип_расхода, Значения: Сумма)."
ws_dash.cell(row=7, column=1).value = "6. Создайте круговую диаграмму для распределения расходов."
ws_dash.cell(row=8, column=1).value = "7. Добавьте срезы: Год, Квартал, Месяц, Материал, Тип_расхода."

# Remove default sheet
wb.remove(wb["Sheet"])

# Save workbook
wb.save("Production_Accounting.xlsx")

print("Excel file 'Production_Accounting.xlsx' created successfully!")